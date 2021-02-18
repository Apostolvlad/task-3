from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side, Color
import requests
from bs4 import BeautifulSoup

FILL_SPLIT = PatternFill(fill_type='solid', fgColor= Color(rgb='cfb53b'))
FILL_TITLE = PatternFill(fill_type='solid', fgColor= Color(rgb='42aaff'))

def set_fill(ws, i, color):
    for y in ws[f'A{i}:G{i}'][0]:
        y.fill = color

def parsing(ws, url):
    soup = BeautifulSoup(requests.get(url).text, features="html.parser")
    h1 = soup.h1.text.replace('\n', '').lstrip().rstrip()

    data_section = soup.find('a', class_ = 'js-popup-open button big invert w250')['data-section']
    data_con = soup.find('div', class_ = 'org-heading-right')['data-con']
    i_add = 0
    for master in soup.find_all('div', class_ = 'executors-block type2'):
        for zakaz in master.find_all('div', class_ = 'org-table-row'):
            #if not zakaz.get('data-f', '').startswith('f-'): continue
            if zakaz.find('div', class_ = 'org-table-col price') is None: continue
            description = zakaz.find('div', class_ = 'org-table-col short-desc')
            if not description is None: description = description.text.replace('\n', '').lstrip().rstrip()
            description_full = zakaz.find('div', class_ = 'org-table-row-hidden-text description full-desc')
            if not description_full is None: description_full = description_full.text.replace('\n', '').lstrip().rstrip()
            
            ws.append([url, h1, description, description_full, data_section, data_con, zakaz.get('data-f')])
            i_add += 1
    print(i_add)
    for _ in range(30 - i_add):
        ws.append([url, h1, '', '', data_section, data_con, zakaz.get('data-f')])
    ws.append([])
    set_fill(ws, ws.max_row + 1, FILL_SPLIT)

#<a href="javascript:void(0)" type="button" class="org-table-row-link active">Положить</a>

#<div class="org-table-col short-desc">Общий массаж тела  </div>

def main():
    with open('urls.txt') as f: base_urls = tuple(f.read().split('\n'))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Результат'
    ws.append(['URL', 'H1', 'Краткое описание', 'Описание', 'ID sec', 'ID cont', 'ID заказа'])
    set_fill(ws, 1, FILL_TITLE)
    for url in base_urls:
        parsing(ws, url)
    wb.save("result.xlsx")

if __name__ == '__main__':
    main()
    